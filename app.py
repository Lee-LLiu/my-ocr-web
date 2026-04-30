import streamlit as st
from aip import AipOcr
from openpyxl import load_workbook
from openpyxl.drawing.image import Image as XLImage
from PIL import Image as PILImage
from fuzzywuzzy import fuzz
import io

# --- 页面配置 ---
st.set_page_config(page_title="价签识别系统", layout="wide")
st.title("果蔬价签识别系统")

with st.sidebar:
    st.header("🔑 百度 API 配置")
    app_id = st.text_input("APP_ID", type="password")
    api_key = st.text_input("API_KEY", type="password")
    secret_key = st.text_input("SECRET_KEY", type="password")
    st.info("💡 提示：本版本会自动过滤水印时间（如 13:50），并将多张同类照片横向填入 Excel。")

# --- 核心匹配引擎 (长词优先) ---
def get_smart_match(full_text, excel_names, alias_dict):
    all_candidates = []
    for name in excel_names:
        all_candidates.append({"std": name, "match": name})
    for std, aliases in alias_dict.items():
        for a in aliases:
            all_candidates.append({"std": std, "match": a})
    
    # 按匹配词长度降序排列，防止“萝卜”截胡“胡萝卜”
    all_candidates.sort(key=lambda x: len(x['match']), reverse=True)

    # 1. 包含匹配
    for item in all_candidates:
        if item['match'] and item['match'] in full_text:
            return item['std']
                
    # 2. 模糊匹配 (兜底)
    best_score = 0
    best_name = "未知"
    for name in excel_names:
        if not name: continue
        score = fuzz.partial_ratio(name, full_text)
        if score > 85 and score > best_score:
            best_score = score
            best_name = name
    
    return best_name if best_score > 85 else "未知"

# --- 价格识别逻辑 (强力过滤水印) ---
def process_ocr_logic(img_bytes, excel_names, alias_dict, client):
    img_for_size = PILImage.open(io.BytesIO(img_bytes))
    img_width, img_height = img_for_size.size
    
    res = client.accurate(img_bytes)
    items = res.get('words_result', [])
    full_text = "".join([item['words'] for item in items])
    
    target_name = get_smart_match(full_text, excel_names, alias_dict)
    
    potential_prices = []
    for item in items:
        text = item['words']
        loc = item['location']
        
        # 【改进 1】过滤掉包含冒号的水印时间 (如 13:50)
        if ":" in text: continue
        
        # 【改进 2】过滤掉包含日期的干扰
        if "-" in text and len(text) > 8: continue

        # 【改进 3】过滤掉底部区域 (提高到 30% 防止水印位置偏高)
        if loc['top'] > img_height * 0.7: continue 

        # 提取纯数字和小数点
        nums = "".join(filter(lambda x: x.isdigit() or x == '.', text))
        
        if len(nums) >= 2:
            # 排除过窄的干扰项（如条形码线条）
            if loc['width'] < 8: continue 
            
            area = loc['width'] * loc['height']
            potential_prices.append({"val": nums, "area": area})
    
    final_price = 0.00
    if potential_prices:
        # 取面积最大的数字作为价格
        best_match = max(potential_prices, key=lambda x: x['area'])['val']
        clean_num = "".join(filter(str.isdigit, best_match))
        
        # 自动补全小数点 (处理 358 -> 3.58 的情况)
        if "." not in best_match and len(clean_num) >= 3:
            final_price = float(int(clean_num)/100)
        else:
            try: final_price = float(best_match)
            except: final_price = 0.00

    return target_name, final_price

# --- 界面交互逻辑 ---
up_template = st.file_uploader("1. 上传 Excel 模板", type=['xlsx'])
up_imgs = st.file_uploader("2. 上传价签照片 (支持多张)", type=['jpg', 'png', 'jpeg'], accept_multiple_files=True)

if st.button("🚀 开始自动化处理"):
    if not (up_template and up_imgs and app_id):
        st.error("请检查：Excel模板、照片、API配置是否完整？")
    else:
        client = AipOcr(app_id, api_key, secret_key)
        wb = load_workbook(io.BytesIO(up_template.read()))
        ws = wb.worksheets[0] # 强制锁定第一个 Sheet
        
        # 提取 Excel A 列关键词
        excel_names = [str(ws.cell(row=i, column=1).value).strip() for i in range(2, ws.max_row + 1) if ws.cell(row=i, column=1).value]
        
        # 提取 Sheet2 别名
        alias_dict = {}
        if len(wb.sheetnames) > 1:
            alias_ws = wb.worksheets[1]
            for row in range(1, alias_ws.max_row + 1):
                std = str(alias_ws.cell(row=row, column=1).value).strip()
                als = str(alias_ws.cell(row=row, column=2).value).strip().split(',')
                if std and als: alias_dict[std] = als

        # 遍历处理每一张图
        for img_file in up_imgs:
            img_bytes = img_file.read()
            name, price = process_ocr_logic(img_bytes, excel_names, alias_dict, client)
            
            if name == "未知":
                st.warning(f"⚠️ {img_file.name}: 未匹配到商品，跳过。")
                continue

            matched = False
            for row in range(2, ws.max_row + 1):
                # 【改进】只要匹配引擎返回的标准名与本行 A 列一致，就进行追加
                if str(ws.cell(row=row, column=1).value).strip() == name:
                    # 自动寻找横向空位 (从 C 列/第 3 列开始，步长为 2)
                    curr_col = 3
                    while ws.cell(row=row, column=curr_col).value is not None:
                        curr_col += 2
                    
                    # 填入价格
                    ws.cell(row=row, column=curr_col + 1).value = price
                    
                    # 处理图片并插入
                    img_pil = PILImage.open(io.BytesIO(img_bytes))
                    if img_pil.mode in ("RGBA", "P"): img_pil = img_pil.convert("RGB")
                    base_width = 800
                    h_size = int((float(img_pil.size[1]) * float(base_width / float(img_pil.size[0]))))
                    img_pil = img_pil.resize((base_width, h_size), PILImage.LANCZOS)
                    img_io = io.BytesIO()
                    img_pil.save(img_io, format="JPEG", quality=85)
                    
                    xl_img = XLImage(img_io)
                    xl_img.width = 90
                    xl_img.height = int(h_size * (90 / base_width))
                    
                    # 设置行高并插入图片到计算出的坐标
                    ws.row_dimensions[row].height = xl_img.height * 0.8
                    coord = ws.cell(row=row, column=curr_col).coordinate
                    ws.add_image(xl_img, coord)
                    
                    st.success(f"✅ {img_file.name} -> 【{name}】已存入第 {curr_col} 列，价格: {price}")
                    matched = True
                    break
            
            if not matched:
                st.warning(f"⚠️ {img_file.name}: 虽识别为 {name} 但未在 Excel 找到对应行。")

        # 保存结果
        out_io = io.BytesIO()
        wb.save(out_io)
        st.download_button("📥 下载识别结果", data=out_io.getvalue(), file_name="result_final.xlsx")
